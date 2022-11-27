import re
import logging
from datetime import datetime

import psycopg
from flask import jsonify
from openpyxl import load_workbook
from db import connection


def import_data():
    '''Импорт данных из файла.
    '''
    with connection() as conn:
        '''Очистка таблицы в БД.
        Если такой таблицы не существует, то она создается.
        '''
        cur = conn.cursor()
        try:
            with conn.transaction():
                cur.execute('TRUNCATE TABLE table1')
        except psycopg.ProgrammingError:
            cur.execute('CREATE TABLE table1 (rep_dt DATE, delta NUMERIC(10, 2))')

        '''Создание представления для таблицы в БД со смещением данных deltalag
        '''
        cur.execute('CREATE OR REPLACE VIEW table1_view(rep_dt, delta) '
                    'AS SELECT rep_dt, delta, LAG(delta, -2) '
                    'OVER (ORDER BY rep_dt) deltalag from table1')

        '''Запись данных в БД из коллекции строк.
        Если строка при записи вызывает исключение, то она сохраняется
        в отдельный файл. После чего переходим к записи следующей строки.
        '''
        sql_insert_row = ("INSERT INTO table1 (Rep_dt, Delta) VALUES (%s, %s)")
        for row in get_rows_from_file():
            try:
                with conn.transaction():
                    '''Вставка строки в таблицу.
                    Если возникнет исключение, то эта транзакция откатится.
                    '''
                    data_row = (parse_date(row[0].value),
                                parse_delta(row[1].value))
                    cur.execute(sql_insert_row, data_row)
            except psycopg.DataError as e:
                '''Строка с данными, которая вызвала исключение,
                сохраняется в отдельный файл.
                '''
                # TODO: file.write(row)
                logging.error('{} {}'.format(str(e), type(e)))

    response = {'result': 0, 'resultStr': 'OK'}
    return jsonify(response)


def get_rows_from_file(fname='testData.xlsx'):
    wb = load_workbook(filename=fname)
    sheet = wb.active
    return sheet.iter_rows(2, sheet.max_row)


def parse_date(value):
    value = str(value)
    if re.match(r'\d{2}[.]\d{2}[.]\d{4}', value):
        return datetime.strptime(value, '%d.%m.%Y')
    elif re.match(r'\d{4}[-]\d{2}[-]\d{2}', value):
        return datetime.strptime(value, '%Y-%m-%d')
    else:
        raise psycopg.DataError('Unexpected date format')


def parse_delta(value):
    return str(value).replace(',', '.')


'''
def parse_delta(value):
    if isinstance(value, (int, float)):
        return value
    elif isinstance(value, str):
        try:
            return float(value.replace(',', '.'))
        except ValueError:
            raise psycopg.DataError('Unexpected data format')
    else:
        raise psycopg.DataError('Unexpected data format')
'''
